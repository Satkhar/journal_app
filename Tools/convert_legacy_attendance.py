#!/usr/bin/env python3
"""Convert the legacy attendance workbook into the importer JSON contract.

This script never writes SQLite directly. Database ownership stays in the C++
storage adapter; the spreadsheet-specific heuristics remain isolated here.
"""

from __future__ import annotations

import argparse
import json
import re
import uuid
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl


RANKS = {
    "оруженосцы": "squire",
    "пажи": "page",
    "рекруты": "recruit",
    "гости (новички)": "guest",
}
HAND = {"правша": "right", "левша": "left"}
PAYMENT_COLORS = {"FFC6DEB5", "FFA9CE91"}
SPECIAL_TRAINING_COLOR = "FFD4A8FF"
FIRST_VISIT_COLORS = {"FF9DC3E6", "FFBDD7EE"}
NO_PAYMENT_COLORS = {"FFFCA4A4", "FFFFC7CE"}
OTHER_COLORS = {"FFF4B183", "FFF8CBAD"}


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def split_name(name: str, rank: str) -> tuple[str, str]:
    match = re.fullmatch(r"(.+?)\s*\(([^()]*)\)\s*", name)
    if rank == "recruit" and match and match.group(2).casefold() == "тверь":
        return "", match.group(1).strip()
    if rank in {"page", "squire", "knight"}:
        if match:
            return match.group(1).strip(), match.group(2).strip()
        return name.strip(), ""
    return "", name.strip()


def rgb(cell: openpyxl.cell.cell.Cell) -> str:
    color = cell.fill.fgColor
    return str(color.rgb).upper() if color.type == "rgb" else ""


def extract_comment(cell: openpyxl.cell.cell.Cell) -> str:
    if not cell.comment:
        return ""
    text = cell.comment.text
    if "Comment:" in text:
        text = text.split("Comment:", 1)[1]
    return " ".join(text.replace("\t", " ").split())


def add_marker(markers: dict[tuple[str, int], dict], participant_id: str,
               day: int, kind: int, note: str = "") -> None:
    marker = markers.setdefault(
        (participant_id, day),
        {"participant_id": participant_id, "day": day,
         "kind_mask": 0, "notes": []},
    )
    marker["kind_mask"] |= kind
    if note and note not in marker["notes"]:
        marker["notes"].append(note)


def convert(input_path: Path, sheet_name: str | None) -> dict:
    workbook = openpyxl.load_workbook(input_path, data_only=False)
    sheet = workbook[sheet_name] if sheet_name else workbook.active
    date_cells = [sheet.cell(2, column) for column in range(5, 14)]
    dates = [cell.value for cell in date_cells]
    if not dates or any(not isinstance(value, datetime) for value in dates):
        raise ValueError("В E2:M2 ожидались даты тренировок")
    year, month = dates[0].year, dates[0].month
    if any(value.year != year or value.month != month for value in dates):
        raise ValueError("Все даты должны относиться к одному месяцу")

    imported_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    participants: list[dict] = []
    attendance: list[dict] = []
    markers: dict[tuple[str, int], dict] = {}
    current_rank = ""
    rows: list[tuple[int, str, str]] = []
    for row in range(3, 52):
        group = clean(sheet.cell(row, 1).value).casefold()
        if group in RANKS:
            current_rank = RANKS[group]
        name = clean(sheet.cell(row, 2).value)
        if name and current_rank:
            rows.append((row, name, current_rank))
    for row in range(55, 62):
        name = clean(sheet.cell(row, 1).value)
        if name:
            rows.append((row, name, "knight"))

    namespace = uuid.uuid5(uuid.NAMESPACE_URL,
                           f"journal-legacy:{sheet.title}:{year}-{month}")
    for row, source_name, rank in rows:
        participant_id = str(uuid.uuid5(namespace, f"row:{row}:{source_name}"))
        historical_name, full_name = split_name(source_name, rank)
        notes: list[str] = []
        payment_note = clean(sheet.cell(row, 15).value)
        if payment_note:
            notes.append(f"[{year:04d}-{month:02d}] Оплата: {payment_note}")
        comment = extract_comment(sheet.cell(row, 14))
        if comment:
            notes.append(f"[{imported_at}] {comment}")
        participants.append({
            "id": participant_id,
            "historical_name": historical_name,
            "full_name": full_name,
            "contact": clean(sheet.cell(row, 4).value),
            "rank": rank,
            "combat_hand": HAND.get(clean(sheet.cell(row, 3).value).casefold(),
                                    "unknown"),
            "notes": "\n".join(notes),
        })

        trauma_days: list[int] = []
        for column, date in zip(range(5, 14), dates):
            cell = sheet.cell(row, column)
            value = clean(cell.value)
            if value == "+":
                attendance.append({"participant_id": participant_id,
                                   "day": date.day})
            elif value.casefold() == "травма":
                trauma_days.append(date.day)
            color = rgb(cell)
            if color in PAYMENT_COLORS:
                add_marker(markers, participant_id, date.day, 1, "Оплата")
            if color == SPECIAL_TRAINING_COLOR:
                add_marker(markers, participant_id, date.day, 2,
                           "Железная тренировка")
            if color in FIRST_VISIT_COLORS:
                add_marker(markers, participant_id, date.day, 4,
                           "Первая тренировка")
            if color in NO_PAYMENT_COLORS:
                add_marker(markers, participant_id, date.day, 8,
                           "Оплаты нет — обратить внимание")
            if color in OTHER_COLORS:
                add_marker(markers, participant_id, date.day, 8,
                           value or "Уточнить")
        if trauma_days:
            day_list = ", ".join(f"{day:02d}.{month:02d}.{year}" for day in
                                 trauma_days)
            add_marker(markers, participant_id, trauma_days[0], 8,
                       f"Травма; даты: {day_list}")

    marker_list = []
    for marker in sorted(markers.values(),
                         key=lambda item: (item["participant_id"], item["day"])):
        marker_list.append({
            "participant_id": marker["participant_id"],
            "day": marker["day"],
            "kind_mask": marker["kind_mask"],
            "note": "; ".join(marker.pop("notes")),
        })
    return {
        "format_version": 1,
        "source_sheet": sheet.title,
        "converted_at": imported_at,
        "year": year,
        "month": month,
        "active_days": [value.day for value in dates],
        "participants": participants,
        "attendance": attendance,
        "day_markers": marker_list,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--sheet")
    args = parser.parse_args()
    result = convert(args.input, args.sheet)
    args.output.write_text(json.dumps(result, ensure_ascii=False, indent=2),
                           encoding="utf-8")
    print(f"Участников: {len(result['participants'])}; "
          f"посещений: {len(result['attendance'])}; "
          f"отметок: {len(result['day_markers'])}")


if __name__ == "__main__":
    main()
