#!/usr/bin/env python3
"""Convert the multi-sheet Arme attendance workbook to import JSON v2.

The workbook mixes a monthly roster, attendance, per-day colour markers and a
month-level payment status.  This converter keeps all source interpretation in
one auditable place; the C++ importer remains responsible for SQLite schema and
transactions.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import tempfile
import uuid
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

import openpyxl


RANKS = {
    "оруженосцы": "squire",
    "пажи": "page",
    "рекруты": "recruit",
    "гости": "guest",
    "новички": "novice",
}
HISTORY_RANKS = {"novice", "recruit", "page", "squire", "knight"}
HAND_MAP = {"правша": "right", "левша": "left", "?": "unknown"}

PAYMENT = 0x01
SPECIAL_TRAINING = 0x02
FIRST_VISIT = 0x04
OTHER = 0x08

LEGEND_LABELS = {
    "первая трена",
    "был(а), оплачено",
    "был(а), нет оплаты",
    "уточнить/см. примечание",
    "железная трена",
    "отсутствовал",
}


@dataclass(frozen=True)
class Occurrence:
    year: int
    month: int
    row: int
    raw_name: str
    rank: str
    hand: str
    contact: str
    knight_block: bool
    sheet_title: str


@dataclass
class ParsedMonth:
    year: int
    month: int
    active_days: list[int]
    occurrences: list[Occurrence]
    rows: dict[int, Occurrence]
    sheet: openpyxl.worksheet.worksheet.Worksheet
    date_columns: dict[int, int]
    payment_column: int | None


@dataclass(frozen=True)
class IdentityRules:
    aliases: dict[str, str]
    recruit_historical_forms: set[str]
    profile_overrides: dict[str, dict[str, str]]


@dataclass
class Marker:
    kind_mask: int = 0
    notes: list[str] = field(default_factory=list)

    def add(self, kind: int, note: str = "") -> None:
        self.kind_mask |= kind
        note = clean(note)
        if note and note not in self.notes:
            self.notes.append(note)


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\\n", " ").split())


def normalize_name(value: str) -> str:
    result = clean(value).casefold().replace("ё", "е")
    result = re.sub(r"\s*\(\s*", " (", result)
    result = re.sub(r"\s*\)\s*", ")", result)
    return result


def load_identity_rules(path: Path) -> IdentityRules:
    raw = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(raw, dict) or not isinstance(
        raw.get("identity_aliases"), list
    ):
        raise ValueError("Некорректный файл identity rules")
    aliases: dict[str, str] = {}
    known_identities: set[str] = set()
    for group in raw["identity_aliases"]:
        if (
            not isinstance(group, dict)
            or not isinstance(group.get("identity"), str)
            or not group["identity"].strip()
            or not isinstance(group.get("names"), list)
            or not group["names"]
        ):
            raise ValueError("Некорректная группа алиасов")
        identity = group["identity"].strip()
        if identity in known_identities:
            raise ValueError(f"Повтор identity: {identity}")
        known_identities.add(identity)
        for name in group["names"]:
            if not isinstance(name, str) or not clean(name):
                raise ValueError(f"Пустой алиас в {identity}")
            normalized = normalize_name(name)
            if normalized in aliases:
                raise ValueError(f"Алиас повторён: {name}")
            aliases[normalized] = identity

    forms_value = raw.get("recruit_historical_forms", [])
    overrides_value = raw.get("profile_overrides", {})
    if not isinstance(forms_value, list) or not isinstance(overrides_value, dict):
        raise ValueError("Некорректные overrides identity rules")
    forms = {
        normalize_name(value)
        for value in forms_value
        if isinstance(value, str) and clean(value)
    }
    if len(forms) != len(forms_value):
        raise ValueError("Некорректная форма имени рекрута")
    overrides: dict[str, dict[str, str]] = {}
    for identity, values in overrides_value.items():
        if not isinstance(identity, str) or not isinstance(values, dict):
            raise ValueError("Некорректный profile override")
        parsed: dict[str, str] = {}
        for key, value in values.items():
            if key not in {"historical_name", "full_name", "notes"} or not isinstance(
                value, str
            ):
                raise ValueError(f"Некорректное поле override: {identity}/{key}")
            parsed[key] = value
        overrides[identity] = parsed
    return IdentityRules(aliases, forms, overrides)


def normalize_contact(value: str) -> str:
    result = clean(value)
    if not result:
        return ""
    if result.startswith("@"):
        return result
    lowered = result.casefold()
    for prefix in ("https://vk.com/", "http://vk.com/", "vk.com/"):
        if lowered.startswith(prefix):
            handle = result[len(prefix) :].strip(" /")
            return f"https://vk.com/{handle}"
    return f"https://vk.com/{result.strip(' /')}"


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def sanitized_workbook_copy(source: Path, target: Path) -> None:
    """Repair four invalid border colours without touching user input."""
    with zipfile.ZipFile(source, "r") as input_zip:
        with zipfile.ZipFile(target, "w", zipfile.ZIP_DEFLATED) as output_zip:
            for info in input_zip.infolist():
                data = input_zip.read(info.filename)
                if info.filename == "xl/styles.xml":
                    data = data.replace(b'rgb="none"', b'rgb="00000000"')
                output_zip.writestr(info, data)


def load_workbook(source: Path) -> openpyxl.Workbook:
    try:
        return openpyxl.load_workbook(source, data_only=False)
    except ValueError as error:
        details = f"{error} {error.__cause__ or ''}"
        if "Colors must be aRGB hex values" not in details:
            raise
    temporary = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temporary_path = Path(temporary.name)
    temporary.close()
    try:
        sanitized_workbook_copy(source, temporary_path)
        return openpyxl.load_workbook(temporary_path, data_only=False)
    finally:
        temporary_path.unlink(missing_ok=True)


def rank_from_header(value: str) -> str | None:
    normalized = normalize_name(value)
    if normalized in RANKS:
        return RANKS[normalized]
    if normalized.startswith("гости") or normalized.startswith("выпускники"):
        return "guest"
    return None


def parse_months(workbook: openpyxl.Workbook) -> list[ParsedMonth]:
    result: list[ParsedMonth] = []
    seen_months: set[tuple[int, int]] = set()
    for sheet in workbook.worksheets:
        if "посещаемость" not in sheet.title.casefold():
            continue
        dated_cells = [
            cell
            for cell in sheet[2]
            if isinstance(cell.value, (date, datetime))
        ]
        if not dated_cells:
            raise ValueError(f"{sheet.title}: в строке 2 нет дат")
        year = dated_cells[0].value.year
        month = dated_cells[0].value.month
        if any(
            cell.value.year != year or cell.value.month != month
            for cell in dated_cells
        ):
            raise ValueError(f"{sheet.title}: даты относятся к разным месяцам")
        key = (year, month)
        if key in seen_months:
            raise ValueError(f"Повтор месяца {year:04d}-{month:02d}")
        seen_months.add(key)
        date_columns = {cell.column: cell.value.day for cell in dated_cells}
        active_days = sorted(date_columns.values())
        if len(active_days) != len(set(active_days)):
            raise ValueError(f"{sheet.title}: повтор даты")

        payment_columns = [
            cell.column
            for cell in sheet[2]
            if clean(cell.value).casefold() == "оплачено"
        ]
        if len(payment_columns) > 1:
            raise ValueError(f"{sheet.title}: несколько колонок «Оплачено»")

        rank = ""
        knight_block = False
        occurrences: list[Occurrence] = []
        rows: dict[int, Occurrence] = {}
        for row in range(3, sheet.max_row + 1):
            value_a = clean(sheet.cell(row, 1).value)
            value_b = clean(sheet.cell(row, 2).value)
            if normalize_name(value_a) == "рыцари":
                knight_block = True
                rank = "knight"
                continue
            header_rank = rank_from_header(value_a)
            if header_rank:
                knight_block = False
                rank = header_rank
            if knight_block:
                if value_a:
                    occurrence = Occurrence(
                        year,
                        month,
                        row,
                        value_a,
                        "knight",
                        "unknown",
                        "",
                        True,
                        sheet.title,
                    )
                    occurrences.append(occurrence)
                    rows[row] = occurrence
                continue
            if value_b:
                if not rank:
                    raise ValueError(
                        f"{sheet.title}!B{row}: имя вне известной группы"
                    )
                raw_hand = normalize_name(clean(sheet.cell(row, 3).value))
                if raw_hand not in HAND_MAP and raw_hand:
                    raise ValueError(
                        f"{sheet.title}!C{row}: неизвестная рука «{raw_hand}»"
                    )
                occurrence = Occurrence(
                    year,
                    month,
                    row,
                    value_b,
                    rank,
                    HAND_MAP.get(raw_hand, "unknown"),
                    normalize_contact(clean(sheet.cell(row, 4).value)),
                    False,
                    sheet.title,
                )
                occurrences.append(occurrence)
                rows[row] = occurrence
        result.append(
            ParsedMonth(
                year,
                month,
                active_days,
                occurrences,
                rows,
                sheet,
                date_columns,
                payment_columns[0] if payment_columns else None,
            )
        )
    return sorted(result, key=lambda value: (value.year, value.month))


def occurrence_identity(
    occurrence: Occurrence,
    contacts_by_name: dict[str, set[str]],
    rules: IdentityRules,
) -> str:
    normalized = normalize_name(occurrence.raw_name)
    if normalized in rules.aliases:
        return rules.aliases[normalized]
    if occurrence.knight_block:
        # Knight identities use a separate namespace unless an explicit alias
        # proves that a main-roster entry and a knight are the same person.
        return f"knight:{normalized}"
    contacts = contacts_by_name.get(normalized, set())
    if len(contacts) == 1:
        return f"contact:{next(iter(contacts)).casefold()}"
    return f"name:{normalized}"


def split_name(
    occurrence: Occurrence, rules: IdentityRules
) -> tuple[str, str]:
    raw = clean(occurrence.raw_name)
    normalized = normalize_name(raw)
    match = re.fullmatch(r"(.+?)\s*\((.+)\)", raw)
    if occurrence.rank in {"page", "squire", "knight"}:
        if match:
            return clean(match.group(1)), clean(match.group(2))
        return raw, ""
    if normalized in rules.recruit_historical_forms and match:
        return clean(match.group(2)), clean(match.group(1))
    if raw.casefold().startswith("новичок ") and not raw.startswith("Новичок ("):
        raw = clean(raw[len("Новичок ") :])
    return "", raw


def candidate_score(value: str, index: int) -> tuple[int, int, int]:
    useful = 0 if value in {"?", "(?)"} else 1
    return useful, len(value), index


def build_profiles(
    months: list[ParsedMonth], rules: IdentityRules, warnings: list[str]
) -> tuple[dict[str, dict[str, object]], dict[Occurrence, str]]:
    contacts_by_name: dict[str, set[str]] = defaultdict(set)
    for month in months:
        for occurrence in month.occurrences:
            if occurrence.contact and not occurrence.knight_block:
                contacts_by_name[normalize_name(occurrence.raw_name)].add(
                    occurrence.contact
                )

    ids: dict[Occurrence, str] = {}
    grouped: dict[str, list[Occurrence]] = defaultdict(list)
    for month in months:
        month_ids: set[str] = set()
        for occurrence in month.occurrences:
            identity = occurrence_identity(occurrence, contacts_by_name, rules)
            participant_id = str(
                uuid.uuid5(
                    uuid.NAMESPACE_URL,
                    f"journal-app:arme-attendance:{identity}",
                )
            )
            if participant_id in month_ids:
                raise ValueError(
                    f"{occurrence.sheet_title}: алиасы объединили две строки «"
                    f"{occurrence.raw_name}»"
                )
            month_ids.add(participant_id)
            ids[occurrence] = participant_id
            grouped[identity].append(occurrence)

    profiles: dict[str, dict[str, object]] = {}
    for identity, occurrences in grouped.items():
        occurrences.sort(key=lambda value: (value.year, value.month, value.row))
        participant_id = ids[occurrences[0]]
        historical_candidates: list[tuple[str, int]] = []
        full_candidates: list[tuple[str, int]] = []
        contacts: list[str] = []
        hands: list[str] = []
        raw_names: list[str] = []
        rank_history: list[str] = []
        for index, occurrence in enumerate(occurrences):
            historical, full = split_name(occurrence, rules)
            if historical:
                historical_candidates.append((historical, index))
            if full:
                full_candidates.append((full, index))
            if occurrence.contact:
                contacts.append(occurrence.contact)
            if occurrence.hand != "unknown":
                hands.append(occurrence.hand)
            if occurrence.raw_name not in raw_names:
                raw_names.append(occurrence.raw_name)
            if (
                occurrence.rank in HISTORY_RANKS
                and occurrence.rank not in rank_history
            ):
                rank_history.append(occurrence.rank)

        historical = (
            max(
                historical_candidates,
                key=lambda item: candidate_score(item[0], item[1]),
            )[0]
            if historical_candidates
            else ""
        )
        full = (
            max(
                full_candidates,
                key=lambda item: candidate_score(item[0], item[1]),
            )[0]
            if full_candidates
            else ""
        )
        contacts_by_key: dict[str, str] = {}
        for contact in contacts:
            contacts_by_key.setdefault(contact.casefold(), contact)
        if len(contacts_by_key) > 1:
            sources = ", ".join(
                f"{occurrence.sheet_title}!{occurrence.row}="
                f"{occurrence.contact}"
                for occurrence in occurrences
                if occurrence.contact
            )
            raise ValueError(
                f"{identity}: алиасы объединили разные контакты: {sources}"
            )
        override = rules.profile_overrides.get(identity, {})
        historical = override.get("historical_name", historical)
        full = override.get("full_name", full)
        if not historical and not full:
            raise ValueError(f"Не удалось получить имя для {identity}")
        if len(set(hands)) > 1:
            warnings.append(
                f"{historical or full}: в книге указаны разные боевые руки; "
                f"выбрана последняя ({hands[-1]})"
            )
        notes = override.get("notes", "")
        profiles[participant_id] = {
            "id": participant_id,
            "historical_name": historical,
            "full_name": full,
            "contact": next(iter(contacts_by_key.values()), ""),
            "rank": occurrences[-1].rank,
            "combat_hand": hands[-1] if hands else "unknown",
            "notes": notes,
            "rank_history": [
                {"rank": rank, "obtained_on": None} for rank in rank_history
            ],
            "source_aliases": raw_names,
        }
    return profiles, ids


def colour_signature(cell: openpyxl.cell.cell.Cell) -> tuple[str, object, float]:
    color = cell.fill.fgColor
    if cell.fill.patternType in {None, "none"}:
        return "none", "", 0.0
    if color.type == "rgb":
        return "rgb", str(color.rgb).upper(), float(color.tint)
    if color.type == "theme":
        return "theme", int(color.theme), float(color.tint)
    return color.type or "none", "", float(color.tint)


def classify_date_colour(cell: openpyxl.cell.cell.Cell) -> tuple[int, str]:
    color_type, value, tint = colour_signature(cell)
    if color_type == "rgb":
        if value in {"FFC6DEB5", "FFA9CE91", "FFA9D18E"}:
            return PAYMENT, ""
        if value == "FFD4A8FF":
            return SPECIAL_TRAINING, ""
        if value == "FF00B0F0":
            return FIRST_VISIT, ""
        if value in {"FFFCA4A4", "FFFFC7CE", "FFFF0000"}:
            return OTHER, "Нет оплаты"
        if value in {"FFF4B183", "FFF8CBAD", "FFFFC000", "FFFF9900"}:
            return OTHER, "Уточнить по исходному журналу"
        if value in {"FFFFFFFF", "00000000"}:
            return 0, ""
    if color_type == "theme":
        if value == 9:
            return PAYMENT, ""
        # The workbook's later legend changed first-visit blue to pale yellow.
        if value == 7 and abs(tint - 0.6) < 0.01:
            return FIRST_VISIT, ""
        if value == 0:
            return 0, ""
    return 0, ""


def classify_payment_status(cell: openpyxl.cell.cell.Cell) -> tuple[int, str]:
    color_type, value, _ = colour_signature(cell)
    if color_type == "rgb":
        if value in {"FFC6DEB5", "FFA9CE91", "FFA9D18E"}:
            return PAYMENT, ""
        if value in {"FFFCA4A4", "FFFFC7CE", "FFFF0000"}:
            return OTHER, "Оплаты нет (сводная отметка Excel)"
        if value in {"FFFFC000", "FFFF9900", "FFF4B183", "FFF8CBAD"}:
            return OTHER, "Уточнить оплату (сводная отметка Excel)"
        if value in {"FFFFFFFF", "00000000"}:
            return 0, ""
    if color_type == "theme":
        if value == 9:
            return PAYMENT, ""
        if value == 7:
            return OTHER, "Уточнить оплату (сводная отметка Excel)"
        if value == 0:
            return 0, ""
    return 0, ""


def marker_for(
    markers: dict[tuple[str, int], Marker], participant_id: str, day: int
) -> Marker:
    return markers.setdefault((participant_id, day), Marker())


def notes_beyond_payment(month: ParsedMonth, row: int) -> list[str]:
    if month.payment_column is None:
        return []
    result: list[str] = []
    normalized_legend = {normalize_name(value) for value in LEGEND_LABELS}
    for column in range(month.payment_column + 1, month.sheet.max_column + 1):
        value = clean(month.sheet.cell(row, column).value)
        if value and normalize_name(value) not in normalized_legend:
            result.append(value)
    return result


def combine_note(existing: str, additional: Iterable[str]) -> str:
    parts = [clean(existing)] if clean(existing) else []
    for value in additional:
        normalized = clean(value)
        if normalized and normalized not in parts:
            parts.append(normalized)
    return "; ".join(parts)


def build_month_contract(
    month: ParsedMonth,
    ids: dict[Occurrence, str],
    warnings: list[str],
) -> dict[str, object]:
    participant_ids = [ids[value] for value in month.occurrences]
    attendance: set[tuple[str, int]] = set()
    markers: dict[tuple[str, int], Marker] = {}
    for occurrence in month.occurrences:
        participant_id = ids[occurrence]
        trauma_days: list[int] = []
        for column, day in month.date_columns.items():
            cell = month.sheet.cell(occurrence.row, column)
            value = clean(cell.value)
            normalized_value = value.casefold()
            if normalized_value in {"+", "+(?)"}:
                attendance.add((participant_id, day))
                if normalized_value == "+(?)":
                    marker_for(markers, participant_id, day).add(
                        OTHER, "Посещение отмечено как «+(?)» в Excel"
                    )
            elif normalized_value == "травма":
                trauma_days.append(day)
            elif normalized_value:
                marker_for(markers, participant_id, day).add(
                    OTHER,
                    f"{value} ({day:02d}.{month.month:02d}.{month.year})",
                )

            kind, note = classify_date_colour(cell)
            if kind:
                marker_for(markers, participant_id, day).add(kind, note)

        if trauma_days:
            note = "Травма: " + ", ".join(
                f"{day:02d}.{month.month:02d}.{month.year}"
                for day in trauma_days
            )
            marker_for(markers, participant_id, trauma_days[0]).add(OTHER, note)

        supplemental_notes = notes_beyond_payment(month, occurrence.row)
        payment_text = (
            clean(month.sheet.cell(occurrence.row, month.payment_column).value)
            if month.payment_column is not None
            else ""
        )
        status_kind, status_note = (
            classify_payment_status(
                month.sheet.cell(occurrence.row, month.payment_column)
            )
            if month.payment_column is not None
            else (0, "")
        )
        existing_payment_days = sorted(
            day
            for (marker_id, day), marker in markers.items()
            if marker_id == participant_id and marker.kind_mask & PAYMENT
        )
        existing_other_days = sorted(
            day
            for (marker_id, day), marker in markers.items()
            if marker_id == participant_id and marker.kind_mask & OTHER
        )
        if status_kind == PAYMENT:
            if existing_payment_days:
                target_day = existing_payment_days[-1]
                marker_for(markers, participant_id, target_day).add(
                    0, payment_text
                )
            else:
                target_day = month.active_days[-1]
                note = combine_note(
                    "Сводная отметка оплаты; точная дата не указана",
                    [payment_text],
                )
                marker_for(markers, participant_id, target_day).add(
                    PAYMENT, note
                )
        elif status_kind == OTHER:
            target_day = month.active_days[-1]
            note = combine_note(status_note, [payment_text])
            marker_for(markers, participant_id, target_day).add(OTHER, note)
        elif payment_text:
            if existing_payment_days:
                target_day = existing_payment_days[-1]
            elif existing_other_days:
                target_day = existing_other_days[0]
            else:
                target_day = month.active_days[-1]
                marker_for(markers, participant_id, target_day).add(OTHER)
            marker_for(markers, participant_id, target_day).add(0, payment_text)

        if supplemental_notes:
            existing_other_days = sorted(
                day
                for (marker_id, day), marker in markers.items()
                if marker_id == participant_id and marker.kind_mask & OTHER
            )
            target_day = (
                existing_other_days[0]
                if existing_other_days
                else month.active_days[-1]
            )
            marker_for(markers, participant_id, target_day).add(
                OTHER, "; ".join(supplemental_notes)
            )

    marker_rows = []
    for (participant_id, day), marker in sorted(
        markers.items(), key=lambda item: (item[0][0], item[0][1])
    ):
        note = "; ".join(marker.notes)
        if not marker.kind_mask:
            warnings.append(
                f"{month.year:04d}-{month.month:02d}: отброшено примечание "
                f"без типа для {participant_id} / {day}"
            )
            continue
        if len(note) > 500:
            raise ValueError(
                f"Слишком длинное примечание {month.year}-{month.month}-{day}"
            )
        marker_rows.append(
            {
                "participant_id": participant_id,
                "day": day,
                "kind_mask": marker.kind_mask,
                "note": note,
            }
        )
    return {
        "year": month.year,
        "month": month.month,
        "active_days": month.active_days,
        "participant_ids": participant_ids,
        "attendance": [
            {"participant_id": participant_id, "day": day}
            for participant_id, day in sorted(attendance)
        ],
        "day_markers": marker_rows,
    }


def report_text(contract: dict[str, object], warnings: list[str]) -> str:
    participants = contract["participants"]
    months = contract["months"]
    aliases = [
        profile
        for profile in participants
        if len(profile.get("source_aliases", [])) > 1
    ]
    marker_kinds: Counter[str] = Counter()
    for month in months:
        for marker in month["day_markers"]:
            mask = marker["kind_mask"]
            for name, flag in (
                ("оплата", PAYMENT),
                ("железная тренировка", SPECIAL_TRAINING),
                ("первая тренировка", FIRST_VISIT),
                ("другое", OTHER),
            ):
                if mask & flag:
                    marker_kinds[name] += 1
    lines = [
        "Отчёт преобразования «посещаемость Армэ»",
        f"SHA-256 источника: {contract['source_sha256']}",
        f"Месяцев: {len(months)}",
        f"Уникальных участников: {len(participants)}",
        f"Записей состава: {sum(len(m['participant_ids']) for m in months)}",
        f"Посещений: {sum(len(m['attendance']) for m in months)}",
        f"Отметок: {sum(len(m['day_markers']) for m in months)}",
        "Типы отметок: "
        + ", ".join(f"{key}={value}" for key, value in marker_kinds.items()),
        "",
        "Объединённые варианты имён:",
    ]
    for profile in sorted(
        aliases,
        key=lambda value: (
            value["historical_name"] or value["full_name"]
        ).casefold(),
    ):
        name = profile["historical_name"] or profile["full_name"]
        lines.append(f"- {name}: {', '.join(profile['source_aliases'])}")
    lines.extend(["", "Предупреждения:"])
    lines.extend(f"- {warning}" for warning in warnings)
    if not warnings:
        lines.append("- нет")
    return "\n".join(lines) + "\n"


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Преобразовать многомесячный журнал Армэ в JSON v2"
    )
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--identity-map", type=Path, required=True)
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()
    if not args.input.is_file():
        parser.error(f"Файл не найден: {args.input}")
    if not args.identity_map.is_file():
        parser.error(f"Файл identity rules не найден: {args.identity_map}")
    if args.output.exists():
        parser.error(f"Выходной JSON уже существует: {args.output}")
    if args.report and args.report.exists():
        parser.error(f"Отчёт уже существует: {args.report}")

    warnings: list[str] = []
    rules = load_identity_rules(args.identity_map)
    workbook = load_workbook(args.input)
    months = parse_months(workbook)
    profiles, ids = build_profiles(months, rules, warnings)
    month_contracts = [
        build_month_contract(month, ids, warnings) for month in months
    ]
    used_ids = {
        participant_id
        for month in month_contracts
        for participant_id in month["participant_ids"]
    }
    if used_ids != set(profiles):
        raise ValueError("Профили и составы месяцев не согласованы")
    contract = {
        "format_version": 2,
        "source_file": args.input.name,
        "source_sha256": sha256(args.input),
        "identity_rules_sha256": sha256(args.identity_map),
        "participants": sorted(profiles.values(), key=lambda value: value["id"]),
        "months": month_contracts,
    }
    args.output.write_text(
        json.dumps(contract, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    if args.report:
        args.report.write_text(report_text(contract, warnings), encoding="utf-8")
    print(report_text(contract, warnings), end="")


if __name__ == "__main__":
    main()
